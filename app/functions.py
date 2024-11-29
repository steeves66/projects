import openpyxl
import os
from django.http import HttpResponse
from django.conf import settings
from django.apps import apps

class ExcelFile:
    def __init__(self, file:str, sheet_1:str, sheet_1_row:int, sheet_1_col:int, model_name:str, sheet_2:str=None, sheet_2_row:int=None, sheet_2_col:int=None):
        self.file = file
        self.sheet_1 = sheet_1
        self.sheet_1_row = sheet_1_row 
        self.sheet_1_col = sheet_1_col
        self.model_name = model_name
        self.sheet_2 = sheet_2
        self.sheet_2_row = sheet_2_row 
        self.sheet_2_col = sheet_2_col

    def get_model(self):
        for app_config in apps.get_app_configs():  # Iterate over all apps
            try:
                # Try to get the model from each app's models
                model = apps.get_model(app_config.label, self.model_name)
                return model
            except LookupError:
            # If model is not found in the current app, continue to the next app
                continue
        print(f"Model '{self.model_name}' not found in any app.")
        return None
    
    def get_sheet_2_fields(self):
        match self.model_name:
            case 'Aoo':
                queryset = self.get_model().objects.values("id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_ppm","ope_res_pme","pub_aao_bomp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_ppm","ope_res_pme","pub_aao_bomp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
            case 'Aor':
                queryset = self.get_model().objects.values("id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_ppm","ope_res_pme","inf_soum_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_ppm","ope_res_pme","inf_soum_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
            case 'Gag':
                queryset = []
                fields = []
            case 'Pso':
                queryset = self.get_model().objects.values("id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_pspm","ope_res_pme","pub_aao_bomp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","num_aoo","imputation","objet","nat_oper","ope_insc_pspm","ope_res_pme","pub_aao_bomp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
            case 'Psl':
                queryset = self.get_model().objects.values("id","ministere","ac","num_ope","imputation","objet","nat_oper","ope_insc_pspm","ope_res_pme","exam_val_dao_dgmp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","num_ope","imputation","objet","nat_oper","ope_insc_pspm","ope_res_pme","exam_val_dao_dgmp_cst","inf_soum_nn_ret_res_cst","conf_cont_mar_cst","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
            case 'Psc':
                queryset = self.get_model().objects.values("id","ministere","ac","num_ope","imputation","objet","nat_oper","ope_insc_pspm","form_dem_cot_conf_mod_elb_dgmp","not_att_soum_ret_cst","form_sel_conf_mod_dgmp","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","num_ope","imputation","objet","nat_oper","ope_insc_pspm","form_dem_cot_conf_mod_elb_dgmp","not_att_soum_ret_cst","form_sel_conf_mod_dgmp","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
            case 'Psd':
                queryset = self.get_model().objects.values("id","ministere","ac","imputation","objet","nat_oper","ope_insc_pspm","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh")
                fields = ["id","ministere","ac","imputation","objet","nat_oper","ope_insc_pspm","ext_exh_arch_doc_exi","ext_exh_arch_doc_exh"]
        return (queryset, fields)

    def create_excel_file(self):
        # Chemin du fichier Excel
        file_path = os.path.join(settings.FILE_FOLDER, self.file)
        try:
            # Essayer d'ouvrir le fichier Excel
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            return HttpResponse("Fichier Excel inexistant.", status=404)        
            
        sheet_proc = workbook[self.sheet_1]  # Sélectionne la première feuille  

        # Récupérer toutes les données du modèle
        model = self.get_model()
        queryset_sh_1 = model.objects.all()

        # Récupérer les champs du modèle (en excluant les champs spécifiques comme ForeignKey)
        fields = [field.name for field in model._meta.get_fields() if not field.is_relation]      
        
        # Insérer les données dans la prémière feuille
        for row_num, obj in enumerate(queryset_sh_1, start=self.sheet_1_row):
            for col_num, field_name in enumerate(fields, start=self.sheet_1_col):
                # Utiliser getattr pour obtenir la valeur du champ
                sheet_proc.cell(row=row_num, column=col_num, value=getattr(obj, field_name))    
        print(workbook)  

        #Sélectionne la deuxième feuille
        # sheet_data = workbook[self.sheet_2]
        sheet_data = None
        queryset_sh_2, fields_2 = self.get_sheet_2_fields()

        if self.sheet_2 in workbook.sheetnames and fields_2:
            sheet_data = workbook[self.sheet_2]
            for row_num, obj in enumerate(queryset_sh_2, start=self.sheet_2_row):
                for col_num, field_name in enumerate(fields_2, start=self.sheet_2_col):
                    # Utiliser getattr pour obtenir la valeur du champ
                    sheet_data.cell(row=row_num, column=col_num, value=obj[field_name]) 
        else:
            pass

        return workbook